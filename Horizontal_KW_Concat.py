
# coding: utf-8

# In[25]:

import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
	import pandas as pd
	import numpy as np
	import xlrd
except:
    install('pandas')
    install('numpy')
    install('xlrd')

import pandas as pd
import numpy as np


# In[26]:


print("The File should have columns 'Keywords' and 'Groups' accordingly.")
WBname = str(input("Please give the name of the workbook.\n"))


# In[27]:


df = pd.read_excel(WBname)
newdict = {}


# In[28]:


for group in df["Groups"].unique():
    newdict[group] = list(df[df.Groups==group]["Keywords"])


# In[29]:


Horizontal = pd.DataFrame.from_dict(newdict, orient = "index").transpose()


# In[30]:


Horizontal.to_excel("Horizontal_"+WBname)


# In[31]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

